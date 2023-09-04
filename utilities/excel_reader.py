import openpyxl


workbook = 'work_book'
sheet = 'shet_name'

def getRowCount():
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook['']
    return sheet.max_row


def getColumnCount():
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook['']
    return sheet.max_column


def getCellData(row,col):
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook['']
    return sheet.cell(row=row,column=col).value


def setCellData(row,col):
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook['']
    sheet.cell(row=row,column=col).value = 4
    workbook.save()